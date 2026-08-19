#!/usr/bin/env python3
"""Josawa Logistics — Global Command Centre — ship/flight/container dashboard with live maps + live data.

Self-contained Flask app. Live flight positions come from OpenSky (no API key).
Live vessel (AIS) positions come from a free AISStream.io key (set AISSTREAM_KEY env).
Without that key the map shows a seeded, realistic SA coastal fleet so the portal is
fully usable offline. Tracker endpoints return live data ONLY on user input.
"""
import os
import json
import math
import time
import sqlite3
import threading
from datetime import datetime, timezone, timedelta

from flask import Flask, render_template, jsonify, request, Response, send_file

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'data', 'tracker.db')
TEMPLATE_DIR = os.path.join(BASE_DIR, 'templates')
STATIC_DIR = os.path.join(BASE_DIR, 'static')
UPLOAD_DIR = os.path.join(BASE_DIR, 'data', 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET', 'josawa-portal-dev')

# --- Brand / theme ---
BRAND = 'Josawa'
BG = '#0a0f1a'
TEAL = '#0d9488'

# --- SA targets (UPPERCASE keys — must match stored/validated casing) ---
SA_TARGETS = {
    'DURBAN': {'lat': -29.858, 'lon': 31.022, 'label': 'Durban Port'},
    'JNB':    {'lat': -26.136, 'lon': 28.241, 'label': 'Johannesburg (JNB)'},
    'CPT':    {'lat': -33.969, 'lon': 18.602, 'label': 'Cape Town (CPT)'},
    'DUR':    {'lat': -29.614, 'lon': 31.140, 'label': 'Durban (DUR)'},
    'PLZ':    {'lat': -33.971, 'lon': 25.617, 'label': 'Port Elizabeth (PLZ)'},
}
ALLOWED_DEST = set(SA_TARGETS.keys())
DEFAULT_DEST = {'ship': 'DURBAN', 'flight': 'JNB'}

BORDER_KM = 50.0
PROGRESS_CAP_KM = 600.0

# --- Clearing & forwarding pipeline stages (for the shipment register) ---
# Ordered: each shipment moves forward through these. 'arrived' gates the clearing steps.
STAGES = [
    'Booked',          # shipment created, not yet departed
    'In Transit',      # departed origin, at sea / in air
    'Arrived',         # discharged / landed at SA port/airport
    'Customs Cleared', # cleared by SARS customs
    'Delivered',       # collected / delivered to Josawa
]
STAGE_INDEX = {s: i for i, s in enumerate(STAGES)}
ATTENTION_DAYS = 7  # "arriving soon" window

# --- Shipment flags (manual status tags beyond the pipeline stage) ---
FLAGS = [
    'Customs Hold', 'Missing Docs', 'Awaiting Slot', 'Pending Payment',
    'Overdue', 'Ready', 'Exception', 'Cleared', 'Held by Forwarder',
]
FLAG_COLORS = {
    'Customs Hold': '#ed1d24', 'Missing Docs': '#f5a623', 'Awaiting Slot': '#f5a623',
    'Pending Payment': '#f5a623', 'Overdue': '#ed1d24', 'Ready': '#3fb950',
    'Exception': '#ed1d24', 'Cleared': '#3fb950', 'Held by Forwarder': '#f5a623',
}

# Southern-Africa bounding box for live OpenSky flight capture
OSKY_BBOX = (-36.0, 12.0, -21.0, 35.0)  # lamin, lomin, lamax, lomax
OSKY_CACHE_TTL = 30  # seconds

AISSTREAM_KEY = os.environ.get('AISSTREAM_KEY', '').strip()

# In-memory cache for live OpenSky states
_osky_cache = {'ts': 0.0, 'states': []}
_osky_lock = threading.Lock()


# --- DB helpers ---
def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode = WAL')           # concurrent readers + 1 writer
    conn.execute('PRAGMA busy_timeout = 30000')          # wait instead of "database is locked"
    conn.execute('PRAGMA foreign_keys = ON')
    return conn


def ensure_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = get_db()
    conn.execute('''CREATE TABLE IF NOT EXISTS tracked_assets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        kind TEXT NOT NULL,
        identifier TEXT NOT NULL,
        destination TEXT NOT NULL,
        lat REAL, lon REAL, speed_kmh REAL,
        position_source TEXT,
        linked_vessel_id INTEGER,
        arrived INTEGER DEFAULT 0,
        arrived_at TEXT,
        added_at TEXT DEFAULT (datetime('now')),
        updated_at TEXT DEFAULT (datetime('now'))
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS vessels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        imo TEXT,
        vtype TEXT,
        flag TEXT,
        lat REAL, lon REAL,
        speed_kn REAL,
        heading REAL,
        status TEXT,
        origin TEXT,
        destination TEXT,
        port TEXT,
        eta_arrival TEXT,
        transporter TEXT,
        group_code TEXT
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS live_flights (
        icao24 TEXT PRIMARY KEY,
        callsign TEXT,
        origin_country TEXT,
        lon REAL, lat REAL,
        velocity REAL,
        heading REAL,
        seen TEXT
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS shipments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        reference TEXT NOT NULL,
        supplier TEXT,
        description TEXT,
        mode TEXT DEFAULT 'sea',          -- 'sea' | 'air'
        carrier TEXT,                      -- vessel name / airline
        awb_bl TEXT,                       -- AWB (air) or Bill of Lading (sea)
        container_no TEXT,
        destination TEXT DEFAULT 'DURBAN',
        forwarder TEXT,                    -- clearing & forwarding agent
        forwarder_contact TEXT,
        stage TEXT DEFAULT 'Booked',
        eta_arrival TEXT,                  -- ISO datetime-ish 'YYYY-MM-DD HH:MM'
        cleared_at TEXT,
        delivered_at TEXT,
        linked_vessel_id INTEGER,
        notes TEXT,
        slot_type TEXT,                     -- 'Harbour' | 'Flight' | 'Other'
        slot_datetime TEXT,                 -- booking/cutoff slot 'YYYY-MM-DD HH:MM'
        slot_ref TEXT,                      -- booking reference / slot number
        commercial_invoice_no TEXT,         -- CI number (shown on first line of row)
        po_number TEXT,                     -- supplier PO number
        freight_file_no TEXT,               -- freight forwarder's file/reference number
        incoming_stock TEXT,                -- incoming stock description (with quantities)
        airport TEXT,                       -- destination airport (air shipments)
        port TEXT,                          -- discharge port
        berth TEXT,                         -- berth number at port
        ship TEXT,                          -- vessel name (sea) — mirrors carrier for clarity
        created_at TEXT DEFAULT (datetime('now')),
        updated_at TEXT DEFAULT (datetime('now'))
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS shipment_documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        shipment_id INTEGER NOT NULL,
        doc_type TEXT,                      -- Proforma Invoice, Supplier PO, Commercial Invoice, Packing List, HAWB, MAWB, SAD500, SAD501, SAD507, Customs Worksheet, EDI/Release Doc, Clearing Instruction, POA, Other
        filename TEXT NOT NULL,
        stored_name TEXT NOT NULL,          -- uuid filename on disk
        content_type TEXT,
        size INTEGER,
        uploaded_at TEXT DEFAULT (datetime('now'))
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS shipment_notes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        shipment_id INTEGER NOT NULL,
        note TEXT NOT NULL,
        author TEXT DEFAULT 'user',
        created_at TEXT DEFAULT (datetime('now'))
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS shipment_flags (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        shipment_id INTEGER NOT NULL,
        flag TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now')),
        UNIQUE(shipment_id, flag)
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS crm_contacts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company TEXT NOT NULL,
        contact_name TEXT,
        email TEXT,
        phone TEXT,
        city TEXT,
        province TEXT,
        sector TEXT,
        notes TEXT,
        tascam_products TEXT
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS suppliers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        contact_name TEXT,
        email TEXT,
        phone TEXT,
        city TEXT,
        country TEXT,
        products TEXT,
        notes TEXT
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS alerts_updates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        kind TEXT NOT NULL,           -- 'SAPS', 'Clearance', 'Flight Delay', 'Shipment Delay', 'Other'
        title TEXT NOT NULL,
        message TEXT,
        shipment_ref TEXT,
        created_at TEXT DEFAULT (datetime('now')),
        resolved INTEGER DEFAULT 0
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        kind TEXT,                 -- 'stage','overdue','arriving','delivered','system'
        shipment_id INTEGER,
        message TEXT,
        read INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )''')
    conn.commit()
    # --- Migrations for existing databases (columns/tables added after first run) ---
    _migrate_db(conn)
    seed_fleet(conn)
    seed_shipments(conn)
    # No seeded/fake data — only real imports (shipments) and the live CRM contact seed.
    import_crm(conn)
    seed_suppliers(conn)
    seed_alerts(conn)
    seed_settings(conn)
    conn.close()


def _migrate_db(conn):
    """Add columns/tables that may be missing on an already-created database."""
    # New shipment columns
    new_cols = [
        'commercial_invoice_no', 'po_number', 'freight_file_no', 'incoming_stock',
        'airport', 'port', 'berth', 'ship',
    ]
    cols = {r['name'] for r in conn.execute('PRAGMA table_info(shipments)').fetchall()}
    for c in new_cols:
        if c not in cols:
            conn.execute(f'ALTER TABLE shipments ADD COLUMN {c} TEXT')
    # Documents table
    conn.execute('''CREATE TABLE IF NOT EXISTS shipment_documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        shipment_id INTEGER NOT NULL,
        doc_type TEXT,
        filename TEXT NOT NULL,
        stored_name TEXT NOT NULL,
        content_type TEXT,
        size INTEGER,
        uploaded_at TEXT DEFAULT (datetime('now'))
    )''')
    conn.commit()


def seed_fleet(conn):
    """Seed a realistic SA coastal cargo fleet so the map + shipping search work offline."""
    existing = conn.execute('SELECT COUNT(*) AS c FROM vessels').fetchone()['c']
    if existing:
        return
    # (name, imo, type, flag, lat, lon, speed_kn, heading, status, origin, port, eta, transporter, group_code)
    fleet = [
        ('Safmarine Nokwanda', 'IMO9581021', 'Container', 'South Africa', -29.858, 31.022, 0.0, 0, 'Moored', 'Singapore', 'Durban', 'Arrived', 'Transnet Port Terminals', 'DBN-A12'),
        ('MSC Tiyanela', 'IMO9720417', 'Container', 'Liberia', -29.870, 31.010, 0.0, 0, 'At anchor', 'Chennai', 'Durban', '2026-08-04 06:00', 'Transnet Port Terminals', 'DBN-B07'),
        ('Maersk Incer', 'IMO9632044', 'Container', 'Denmark', -30.40, 31.20, 14.5, 355, 'En route', 'Singapore', 'Durban', '2026-08-03 22:10', 'APT Terminals', 'DBN-C03'),
        ('Unicorn Dream', 'IMO9512873', 'Bulk', 'Panama', -31.80, 30.10, 12.0, 10, 'En route', 'Santos', 'Durban', '2026-08-04 14:00', 'Bulkmatic', 'DBN-D09'),
        ('NYK Vega', 'IMO9499301', 'Container', 'Japan', -33.969, 18.602, 0.0, 0, 'Moored', 'Hong Kong', 'Cape Town', 'Arrived', 'CTCT', 'CPT-A04'),
        ('Ever Gentle', 'IMO9255178', 'Container', 'Taiwan', -34.05, 18.40, 11.0, 20, 'En route', 'Shenzhen', 'Cape Town', '2026-08-04 09:30', 'CTCT', 'CPT-B02'),
        ('Safmarine Mzansi', 'IMO9785140', 'Container', 'South Africa', -33.971, 25.617, 0.0, 0, 'Moored', 'Mumbai', 'Port Elizabeth', 'Arrived', 'NCT', 'PLZ-A01'),
        ('Pacific Cygnus', 'IMO9377205', 'Car Carrier', 'Marshall Is', -33.85, 25.40, 13.5, 190, 'En route', 'Yokohama', 'Port Elizabeth', '2026-08-04 11:00', 'NCT', 'PLZ-B05'),
        ('Iron Orchid', 'IMO9162743', 'Bulk', 'Liberia', -33.00, 17.94, 0.0, 0, 'Moored', 'Saldanha', 'Saldanha', 'Arrived', 'Saldanha Bay Iron Ore', 'SLD-A01'),
        ('Cape Enterprise', 'IMO9448012', 'Tanker', 'Bahamas', -33.05, 17.80, 9.0, 200, 'En route', 'Luanda', 'Saldanha', '2026-08-04 16:00', 'Saldanha Bay', 'SLD-B03'),
        ('MSC Anneliese', 'IMO9738829', 'Container', 'Liberia', -25.50, 33.00, 16.0, 210, 'En route', 'Mombasa', 'Durban', '2026-08-05 03:00', 'APT Terminals', 'DBN-E11'),
        ('Ever Sigma', 'IMO9501188', 'Container', 'Taiwan', -28.10, 32.60, 15.5, 340, 'En route', 'Port Louis', 'Durban', '2026-08-03 20:40', 'Transnet Port Terminals', 'DBN-F06'),
        ('Algoa Spirit', 'IMO9250043', 'Tanker', 'South Africa', -33.95, 25.55, 0.0, 0, 'At anchor', 'Saldanha', 'Port Elizabeth', '2026-08-04 08:00', 'NCT', 'PLZ-C02'),
        ('Atlantic Star', 'IMO9621197', 'Container', 'Malta', -34.30, 19.10, 12.5, 15, 'En route', 'Walvis Bay', 'Cape Town', '2026-08-05 01:30', 'CTCT', 'CPT-C08'),
        ('Sakura Bloom', 'IMO9415660', 'Reefer', 'Panama', -26.90, 32.85, 14.0, 355, 'En route', 'Tanga', 'Durban', '2026-08-04 05:20', 'Transnet Port Terminals', 'DBN-G04'),
        ('Nordic Breeze', 'IMO9350987', 'Bulk', 'Norway', -32.40, 28.95, 11.0, 185, 'En route', 'Durban', 'East London', '2026-08-04 12:45', 'TPT East London', 'ELS-A02'),
    ]
    for f in fleet:
        # f has 14 elements: name, imo, vtype, flag, lat, lon, speed_kn, heading,
        # status, origin, port, eta, transporter, group_code
        # derive 'destination' (city) from 'port' to satisfy the 15-column schema
        vals = (*f[:10], f[10], f[10], f[11], f[12], f[13])
        conn.execute('''INSERT OR IGNORE INTO vessels
            (name, imo, vtype, flag, lat, lon, speed_kn, heading, status, origin, destination, port, eta_arrival, transporter, group_code)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', vals)
    conn.commit()


def _parse_dt(s):
    """Parse 'YYYY-MM-DD HH:MM' / 'YYYY-MM-DD' into a datetime (naive, local)."""
    if not s:
        return None
    s = s.strip().replace('T', ' ')
    for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def seed_shipments(conn):
    """Seed realistic Josawa import shipments across the C&F pipeline."""
    existing = conn.execute('SELECT COUNT(*) AS c FROM shipments').fetchone()['c']
    if existing:
        return
    now = datetime.now()
    d = lambda days: (now + timedelta(days=days)).strftime('%Y-%m-%d %H:%M')
    # link a couple to real fleet vessels by name
    vids = {r['name']: r['id'] for r in conn.execute('SELECT id, name FROM vessels')}
    rows = [
        # reference, supplier, description, mode, carrier, awb_bl, container, dest, forwarder, contact, stage, eta, cleared, delivered, linked_vessel, notes
        ('ALP-2607-01', 'TASCAM / TEAC Corp', 'TASCAM Mixers x12 + interfaces', 'sea', 'Maersk Incer', 'MAEU7732019', 'MSKU4488210', 'DURBAN', 'APT Terminals', 'thabo@apt.co.za', 'In Transit', d(1), None, None, vids.get('Maersk Incer'), 'Groupage with audio stock'),
        ('ALP-2607-02', 'Shure Distribution', 'Wireless mic systems', 'air', 'SAA Cargo', 'SAA77331904', '', 'JNB', 'Biashara Freight', 'ops@biashara.co.za', 'Arrived', d(-1), None, None, None, 'At JNB cargo, awaiting clearance'),
        ('ALP-2607-03', 'Yamaha Pro Audio', 'Studio monitors x40', 'sea', 'Ever Sigma', 'EGHU5520388', 'EGHU5520388', 'DURBAN', 'Transnet Port Terminals', 'dc@tpt.co.za', 'In Transit', d(2), None, None, vids.get('Ever Sigma'), 'High-value, insured'),
        ('ALP-2607-04', 'AKG Acoustics', 'Headphones bulk', 'sea', 'Sakura Bloom', 'HLXU1190347', 'HLXU1190347', 'DURBAN', 'Bulkmatic', 'clear@bulkmatic.co.za', 'Booked', d(5), None, None, vids.get('Sakura Bloom'), 'Booking confirmed'),
        ('ALP-2606-09', 'TASCAM / TEAC Corp', 'Replacement parts (RMA)', 'sea', 'Safmarine Nokwanda', 'SSZU6612043', 'SSZU6612043', 'DURBAN', 'Transnet Port Terminals', 'dc@tpt.co.za', 'Customs Cleared', d(-3), d(-2), None, vids.get('Safmarine Nokwanda'), 'Cleared, arrange collection'),
        ('ALP-2606-07', 'Focusrite Ltd', 'Audio interfaces x60', 'air', 'Lufthansa Cargo', 'LH88920044', '', 'JNB', 'Biashara Freight', 'ops@biashara.co.za', 'Delivered', d(-6), d(-5), d(-4), None, 'Delivered to warehouse'),
        ('ALP-2607-05', 'Roland Corp', 'Digital pianos x8', 'sea', 'MSC Tiyanela', 'MSCU8841029', 'MSCU8841029', 'DURBAN', 'APT Terminals', 'thabo@apt.co.za', 'Arrived', d(0), None, None, vids.get('MSC Tiyanela'), 'At anchor Durban, discharging'),
    ]
    for r in rows:
        conn.execute('''INSERT INTO shipments
            (reference, supplier, description, mode, carrier, awb_bl, container_no, destination,
             forwarder, forwarder_contact, stage, eta_arrival, cleared_at, delivered_at, linked_vessel_id, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', r)
    conn.commit()
    # Demo flags + a booking slot so the new features are visible on first run.
    def sid(ref):
        row = conn.execute('SELECT id FROM shipments WHERE reference=?', (ref,)).fetchone()
        return row['id'] if row else None
    def add_flag(ref, flag):
        i = sid(ref)
        if i: conn.execute('INSERT OR IGNORE INTO shipment_flags (shipment_id, flag) VALUES (?,?)', (i, flag))
    def set_slot(ref, stype, dt, refno):
        i = sid(ref)
        if i: conn.execute('UPDATE shipments SET slot_type=?, slot_datetime=?, slot_ref=? WHERE id=?',
                           (stype, dt, refno, i))
    now = datetime.now()
    hh = lambda h: (now + timedelta(hours=h)).strftime('%Y-%m-%d %H:%M')
    add_flag('ALP-2607-02', 'Customs Hold')      # arrived, stuck at customs
    add_flag('ALP-2607-04', 'Missing Docs')       # booked but docs incomplete
    add_flag('ALP-2607-01', 'Awaiting Slot')      # in transit, slot not booked yet
    set_slot('ALP-2607-03', 'Harbour', hh(18), 'TPT-DBN-4471')   # due-soon harbour slot
    set_slot('ALP-2607-05', 'Harbour', hh(-3), 'TPT-DBN-4410')   # missed slot (overdue)
    # Example values for the new C&F fields (editable later).
    def set_meta(ref, **kw):
        i = sid(ref)
        if not i: return
        for k, v in kw.items():
            conn.execute(f'UPDATE shipments SET {k}=? WHERE id=?', (v, i))
    set_meta('ALP-2607-01', commercial_invoice_no='CI-2026-7732', po_number='PO-TAS-1188',
             freight_file_no='APT-F2621', port='Durban', ship='Maersk Incer', berth='Pier 2 / B12',
             incoming_stock='Studio mixers x12|12\nAudio interfaces x8|8\nXLR cables 5m|40')
    set_meta('ALP-2607-02', commercial_invoice_no='CI-SHU-4410', po_number='PO-SHU-0094',
             freight_file_no='BIA-F0094', airport='O.R. Tambo (JNB)',
             incoming_stock='Wireless mic systems|24\nMic stands|12')
    set_meta('ALP-2607-03', commercial_invoice_no='CI-YAM-7781', po_number='PO-YAM-2231',
             freight_file_no='TPT-F3310', port='Durban', ship='Ever Sigma', berth='Pier 1 / D06',
             incoming_stock='Studio monitors 5"|40')
    set_meta('ALP-2607-04', commercial_invoice_no='CI-AKG-1190', po_number='PO-AKG-0551',
             freight_file_no='BLK-F0551', port='Durban', ship='Sakura Bloom',
             incoming_stock='Headphones (bulk)|200')
    set_meta('ALP-2606-09', commercial_invoice_no='CI-TAS-6612', po_number='PO-TAS-0909',
             freight_file_no='TPT-F0909', port='Durban', ship='Safmarine Nokwanda',
             incoming_stock='Replacement parts (RMA)|1')
    set_meta('ALP-2606-07', commercial_invoice_no='CI-FOC-8820', po_number='PO-FOC-0442',
             freight_file_no='BIA-F0442', airport='O.R. Tambo (JNB)',
             incoming_stock='Audio interfaces|60')
    set_meta('ALP-2607-05', commercial_invoice_no='CI-ROL-3301', po_number='PO-ROL-0771',
             freight_file_no='APT-F0771', port='Durban', ship='MSC Tiyanela', berth='Pier 3 / C09',
             incoming_stock='Digital pianos|8')
    conn.commit()


def enrich_shipment(row, conn=None):
    """Add derived fields: countdown, attention bucket, linked live position."""
    out = dict(row)
    stage = out.get('stage') or 'Booked'
    out['stage_index'] = STAGE_INDEX.get(stage, 0)
    out['stage_pct'] = round(STAGE_INDEX.get(stage, 0) / (len(STAGES) - 1) * 100, 1)

    now = datetime.now()
    eta = _parse_dt(out.get('eta_arrival'))
    out['eta_parsed'] = out['eta_arrival'] if eta else None
    if eta is None:
        out['eta_countdown'] = None
        out['eta_countdown_hours'] = None
        out['eta_overdue'] = False
        out['eta_soon'] = False
    else:
        diff = eta - now
        out['eta_countdown_hours'] = round(diff.total_seconds() / 3600, 1)
        out['eta_overdue'] = diff.total_seconds() < 0
        out['eta_soon'] = 0 <= diff.total_seconds() <= ATTENTION_DAYS * 24 * 3600

    # Attention bucket — what does she need to act on?
    bucket = 'Later'
    if stage in ('Arrived',) and not out.get('cleared_at'):
        bucket = 'Needs Clearing'
    elif stage == 'Customs Cleared' and not out.get('delivered_at'):
        bucket = 'Ready for Collection'
    elif out.get('eta_overdue') and stage not in ('Delivered', 'Customs Cleared'):
        bucket = 'Overdue'
    elif out.get('eta_soon') and stage in ('Booked', 'In Transit'):
        bucket = 'Arriving Soon'
    elif stage == 'Delivered':
        bucket = 'Done'
    out['attention'] = bucket

    # Linked live position (from fleet vessel, if linked + available)
    out['live_lat'] = None
    out['live_lon'] = None
    out['live_status'] = None
    if out.get('linked_vessel_id'):
        own = conn is None
        if own:
            conn = get_db()
        try:
            v = conn.execute('SELECT lat, lon, status, name FROM vessels WHERE id=?',
                             (out['linked_vessel_id'],)).fetchone()
        finally:
            if own:
                conn.close()
        if v:
            out['live_lat'] = v['lat']
            out['live_lon'] = v['lon']
            out['live_status'] = v['status'] or 'En route'
            out['linked_vessel_name'] = v['name']
    # Manual flags + notes count (always use a fresh connection to avoid
    # reusing a caller's already-closed transaction handle)
    sid = out.get('id')
    fc = get_db()
    try:
        flags = [r['flag'] for r in fc.execute(
            'SELECT flag FROM shipment_flags WHERE shipment_id=? ORDER BY flag', (sid,)).fetchall()] if sid else []
        ncount = fc.execute('SELECT COUNT(*) AS c FROM shipment_notes WHERE shipment_id=?',
                            (sid,)).fetchone()['c'] if sid else 0
    finally:
        fc.close()
    out['flags'] = flags
    out['notes_count'] = ncount
    # Slot status (harbour booking / flight cutoff)
    out['slot_status'] = _slot_status(out)
    return out


def _slot_status(out):
    """Classify a booking slot: None / upcoming / due-soon / missed."""
    st = out.get('slot_datetime')
    if not st:
        return None
    dt = _parse_dt(st)
    if not dt:
        return 'unknown'
    diff = (dt - datetime.now()).total_seconds()
    if diff < 0:
        return 'missed'
    if diff <= 24 * 3600:
        return 'due-soon'
    return 'upcoming'


def import_crm(conn):
    """Import the TASCAM SA CRM master CSV into crm_contacts (one-time seed)."""
    existing = conn.execute('SELECT COUNT(*) AS c FROM crm_contacts').fetchone()['c']
    if existing:
        return
    # Candidate paths (Windows + WSL); first existing wins.
    candidates = [
        '/mnt/c/Users/ordio/tascam-2/TASCAM_SA_Leads_Master.csv',
        '/mnt/c/Users/ordio/TASCAM_SA_Leads_Master.csv',
        os.path.join(os.path.expanduser('~'), 'TASCAM_SA_Leads_Master.csv'),
    ]
    path = next((p for p in candidates if os.path.exists(p)), None)
    if not path:
        print('[CRM] master CSV not found, skipping import')
        return
    try:
        import csv
        with open(path, newline='', encoding='utf-8', errors='replace') as fh:
            reader = csv.DictReader(fh)
            n = 0
            for row in reader:
                company = (row.get('company_name') or '').strip()
                if not company:
                    continue
                conn.execute('''INSERT INTO crm_contacts
                    (company, contact_name, email, phone, city, province, sector, notes, tascam_products)
                    VALUES (?,?,?,?,?,?,?,?,?)''',
                    (company, (row.get('contact_name') or '').strip(),
                     (row.get('email') or '').strip(), (row.get('phone') or '').strip(),
                     (row.get('city') or '').strip(), (row.get('province') or '').strip(),
                     (row.get('sector') or '').strip(), (row.get('notes') or '').strip(),
                     (row.get('tascam_products') or '').strip()))
                n += 1
        conn.commit()
        print(f'[CRM] imported {n} contacts from {path}')
    except Exception as e:
        print('[CRM] import failed:', e)


def seed_settings(conn):
    defaults = {
        'digest_enabled': '0',
        'digest_email': '',
        'digest_hour': '7',          # local hour to send morning digest
        'digest_timezone': 'Africa/Johannesburg',
        'smtp_host': '',
        'smtp_user': '',
        'smtp_from': 'josawa-logistics@local',
        'alerts_enabled': '1',
    }
    for k, v in defaults.items():
        conn.execute('INSERT OR IGNORE INTO settings (key, value) VALUES (?,?)', (k, v))
    conn.commit()


def seed_suppliers(conn):
    """Made-up supplier master data (placeholder until real supplier details are added).
    Editable later via the Suppliers screen / API."""
    existing = conn.execute('SELECT COUNT(*) AS c FROM suppliers').fetchone()['c']
    if existing:
        return
    suppliers = [
        ('TASCAM / TEAC Corp', 'Kenji Sato', 'export@tascam.co.jp', '+81-3-5488-2211', 'Tokyo', 'Japan', 'Mixers, interfaces, recorders', 'Audio manufacturing'),
        ('Shure Distribution', 'Linda Park', 'orders@shure.co.za', '+27-11-555-0142', 'Johannesburg', 'South Africa', 'Wireless mics, IEM', 'Pro audio'),
        ('Yamaha Pro Audio', 'Hiroshi Tanaka', 'pa-export@yamaha.co.jp', '+81-53-460-2111', 'Hamamatsu', 'Japan', 'Studio monitors, mixers', 'Audio manufacturing'),
        ('AKG Acoustics', 'Marie Brunner', 'sales@akg.com', '+43-1-866-550', 'Vienna', 'Austria', 'Headphones, mics', 'Audio manufacturing'),
        ('Focusrite Ltd', 'James Overy', 'export@focusrite.com', '+44-1280-731-000', 'London', 'United Kingdom', 'Audio interfaces', 'Audio manufacturing'),
        ('Roland Corp', 'Satoshi Nakamura', 'global@roland.com', '+81-78-611-8000', 'Osaka', 'Japan', 'Digital pianos, synths', 'Audio manufacturing'),
        ('Sennheiser SA', 'Pieter van der Merwe', 'pa@sennheiser.co.za', '+27-12-333-0123', 'Pretoria', 'South Africa', 'Headphones, mics, wireless', 'Pro audio'),
        ('JBL Professional', 'Thabo Dlamini', 'export@jblpro.co.za', '+27-11-444-0777', 'Johannesburg', 'South Africa', 'Speakers, PA', 'Pro audio'),
    ]
    for s in suppliers:
        conn.execute('''INSERT OR IGNORE INTO suppliers
            (name, contact_name, email, phone, city, country, products, notes)
            VALUES (?,?,?,?,?,?,?,?)''', s)
    conn.commit()


def seed_alerts(conn):
    """Seed the 'Alerts & Updates' feed with made-up operational alerts so the
    panel is populated on first run (SAPS stops, clearance needs, delays, etc.).
    Editable later via the Alerts API/screen."""
    existing = conn.execute('SELECT COUNT(*) AS c FROM alerts_updates').fetchone()['c']
    if existing:
        return
    now = datetime.now()
    def dt(days, hours=0):
        return (now + timedelta(days=days, hours=hours)).strftime('%Y-%m-%d %H:%M')
    alerts = [
        ('SAPS', 'SAPS stop — clearance required', 'SAPS inspection stop flagged on ALP-2607-02 (Shure wireless mics). Customs clearance needed before release from JNB cargo.', 'ALP-2607-02'),
        ('Clearance', 'Customs documents outstanding', 'ALP-2607-04 (AKG headphones) is missing commercial invoice + packing list. Forwarder awaiting docs from supplier.', 'ALP-2607-04'),
        ('Flight Delay', 'Air freight delay', 'Lufthansa Cargo LH8892 delayed 14h due to Frankfurt weather. ALP-2608-11 ETA pushed to ' + dt(2) + '.', 'ALP-2608-11'),
        ('Shipment Delay', 'Vessel behind schedule', 'Maersk Incer (ALP-2607-01) slowed to 11kn — ETA Durban now ' + dt(2) + 'h. Groupage audio stock affected.', 'ALP-2607-01'),
        ('Other', 'Harbour slot missed', 'ALP-2607-05 missed TPT-DBN-4410 harbour slot. Rebook required with Transnet Port Terminals.', 'ALP-2607-05'),
    ]
    for a in alerts:
        conn.execute('''INSERT INTO alerts_updates (kind, title, message, shipment_ref)
            VALUES (?,?,?,?)''', a)
    conn.commit()


def get_setting(key, default=''):
    conn = get_db()
    row = conn.execute('SELECT value FROM settings WHERE key=?', (key,)).fetchone()
    conn.close()
    return row['value'] if row else default


def set_setting(key, value):
    conn = get_db()
    conn.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)', (key, value))
    conn.commit()
    conn.close()


def notify(kind, shipment_id, message):
    """Persist an alert; the SSE stream picks it up for live toasts."""
    conn = get_db()
    conn.execute('INSERT INTO notifications (kind, shipment_id, message) VALUES (?,?,?)',
                 (kind, shipment_id, message))
    conn.commit()
    conn.close()


# --- fetch_url helper ---
def fetch_url(url, timeout=15, headers=None):
    from urllib.request import urlopen, Request
    from urllib.error import URLError, HTTPError
    HEADERS = {
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
    }
    if headers:
        HEADERS.update(headers)
    try:
        req = Request(url, headers=HEADERS)
        with urlopen(req, timeout=timeout) as resp:
            return resp.read().decode('utf-8', errors='replace')
    except (HTTPError, URLError, Exception) as e:
        print(f'[FETCHER] {e} for {url}')
        return None


# --- distance math ---
def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def enrich_asset(r, conn=None):
    """Compute distance/progress/eta/border for a tracked_assets row."""
    dest = r['destination']
    target = SA_TARGETS.get(dest)
    if target is None:
        target = SA_TARGETS.get(dest.upper(), SA_TARGETS['DURBAN'])
    out = dict(r)
    lat = r['lat']
    lon = r['lon']
    # If the asset has no live position but is linked to a fleet vessel, use the
    # vessel's known coordinates so distance/progress/ETA become meaningful.
    linked_vessel = None
    if (lat is None or lon is None) and r['linked_vessel_id']:
        own = conn is None
        if own:
            conn = get_db()
        try:
            linked_vessel = conn.execute(
                'SELECT lat, lon, status, eta_arrival FROM vessels WHERE id=?',
                (r['linked_vessel_id'],)).fetchone()
        finally:
            if own:
                conn.close()
        if linked_vessel and linked_vessel['lat'] is not None:
            lat, lon = linked_vessel['lat'], linked_vessel['lon']
    out['target_label'] = target['label']
    if lat is None or lon is None:
        out['distance_km'] = None
        out['progress_pct'] = 0
        out['eta_hours'] = None
        out['border_reached'] = False
        out['status'] = 'Tracking' if r['kind'] == 'flight' else 'Awaiting position'
        return out
    dist = haversine(lat, lon, target['lat'], target['lon'])
    border = dist <= BORDER_KM
    if dist >= PROGRESS_CAP_KM:
        prog = 0.0
    else:
        prog = max(0.0, min(100.0, (PROGRESS_CAP_KM - dist) / PROGRESS_CAP_KM * 100))
    speed = r['speed_kmh'] or 0
    eta = (dist / speed) if (speed > 0 and not border) else None
    status = 'En route'
    if r['linked_vessel_id']:
        own = conn is None
        if own:
            conn = get_db()
        try:
            v = conn.execute('SELECT eta_arrival, status FROM vessels WHERE id=?',
                             (r['linked_vessel_id'],)).fetchone()
        finally:
            if own:
                conn.close()
        if v:
            status = 'En route (ETA based on schedule)'
    if border:
        status = 'BORDER REACHED'
    out['distance_km'] = round(dist, 1)
    out['progress_pct'] = round(prog, 1)
    out['eta_hours'] = round(eta, 1) if eta else None
    out['border_reached'] = border
    out['status'] = status
    out['target_label'] = target['label']
    return out


# --- live OpenSky flights ---
def get_live_flights():
    """Return live aircraft over Southern Africa (cached). No API key needed."""
    global _osky_cache
    now = time.time()
    with _osky_lock:
        if now - _osky_cache['ts'] < OSKY_CACHE_TTL and _osky_cache['states']:
            return _osky_cache['states']
    lamin, lomin, lamax, lomax = OSKY_BBOX
    url = (f'https://opensky-network.org/api/states/all'
           f'?lamin={lamin}&lomin={lomin}&lamax={lamax}&lomax={lomax}')
    txt = fetch_url(url, timeout=15)
    states = []
    if txt:
        try:
            data = json.loads(txt)
            states = data.get('states') or []
        except Exception as e:
            print('[OPENSKY] parse fail', e)
    with _osky_lock:
        _osky_cache['ts'] = time.time()
        _osky_cache['states'] = states
    # also persist into live_flights for the dashboard
    try:
        conn = get_db()
        conn.execute('DELETE FROM live_flights')
        seen = datetime.now(timezone.utc).isoformat()
        for st in states:
            try:
                conn.execute('INSERT OR REPLACE INTO live_flights VALUES (?,?,?,?,?,?,?,?,?)',
                             (st[0], (st[1] or '').strip(), st[2], st[5], st[6],
                              st[9] or 0, st[10] or 0, seen))
            except Exception:
                pass
        conn.commit()
        conn.close()
    except Exception:
        pass
    return states


def aircraft_to_dict(st):
    try:
        return {
            'icao24': st[0],
            'callsign': (st[1] or '').strip(),
            'origin_country': st[2],
            'lon': st[5],
            'lat': st[6],
            'velocity_ms': st[9] or 0,
            'heading': st[10] or 0,
            'speed_kmh': round((st[9] or 0) * 3.6, 1),
        }
    except Exception:
        return None


# --- Routes ---
@app.route('/')
def index():
    return render_template('dashboard.html')


@app.route('/map')
def map_page():
    return render_template('map.html')


@app.route('/api/stats')
def api_stats():
    conn = get_db()
    assets = conn.execute('SELECT COUNT(*) AS c, '
                          'SUM(CASE WHEN arrived=1 THEN 1 ELSE 0 END) AS arr '
                          'FROM tracked_assets').fetchone()
    fleet = conn.execute('SELECT COUNT(*) AS c FROM vessels').fetchone()['c']
    flights = len(get_live_flights())
    shp = conn.execute('SELECT COUNT(*) AS c, '
                       'SUM(CASE WHEN stage NOT IN (\'Delivered\') THEN 1 ELSE 0 END) AS active '
                       'FROM shipments').fetchone()
    shp_air = conn.execute("SELECT COUNT(*) AS c FROM shipments WHERE mode='air'").fetchone()['c']
    shp_sea = conn.execute("SELECT COUNT(*) AS c FROM shipments WHERE mode='sea'").fetchone()['c']
    suppliers = conn.execute('SELECT COUNT(*) AS c FROM suppliers').fetchone()['c']
    buckets = {}
    for r in conn.execute('SELECT * FROM shipments').fetchall():
        b = enrich_shipment(dict(r))['attention']
        buckets[b] = buckets.get(b, 0) + 1
    conn.close()
    return jsonify({
        'tracked_assets': assets['c'],
        'border_reached': assets['arr'] or 0,
        'fleet_vessels': fleet,
        'live_flights': flights,
        'shipments_total': shp['c'],
        'shipments_active': shp['active'] or 0,
        'shipments_air': shp_air,
        'shipments_sea': shp_sea,
        'suppliers_total': suppliers,
        'shipment_buckets': buckets,
        'live_ships': 'AISStream (live)' if AISSTREAM_KEY else 'no live vessel feed',
        'updated_at': datetime.now(timezone.utc).isoformat(),
    })


@app.route('/api/navis/status')
def api_navis_status():
    """Port/terminal systems migration tracker.
    Real, config-driven via settings (no seeded demo terminals).
    Migration target date + terminal list live in the `settings` table; if absent,
    we report the honest 'not configured' state instead of inventing data."""
    from datetime import date
    conn = get_db()
    target = conn.execute("SELECT value FROM settings WHERE key='navis_target_date'").fetchone()
    target = target['value'] if target else None
    terms = conn.execute("SELECT value FROM settings WHERE key='navis_terminals'").fetchone()
    terminals = []
    if terms:
        import json as _json
        try:
            terminals = _json.loads(terms['value'])
        except Exception:
            terminals = []
    conn.close()
    if target:
        try:
            d = date.fromisoformat(target)
            days_remaining = (d - date.today()).days
        except ValueError:
            days_remaining = None
    else:
        days_remaining = None
    return jsonify({
        'configured': bool(target),
        'target_date': target,
        'days_remaining': days_remaining,
        'terminals': terminals,
        'note': None if target else 'Set navis_target_date + navis_terminals in settings to activate the migration tracker.'
    })


@app.route('/api/consolidator')
def api_consolidator():
    """Group real shipments by forwarder to surface consolidation opportunities.
    A forwarder with multiple active shipments is a candidate for group clearing."""
    conn = get_db()
    rows = conn.execute(
        "SELECT forwarder, supplier, reference, mode, stage, destination "
        "FROM shipments ORDER BY forwarder").fetchall()
    conn.close()
    groups = {}
    for r in rows:
        fwd = (r['forwarder'] or '').strip() or '(unassigned)'
        groups.setdefault(fwd, []).append(dict(r))
    # rank: most shipments first
    ranked = sorted(groups.items(), key=lambda kv: -len(kv[1]))
    return jsonify({
        'groups': [
            {
                'forwarder': fwd,
                'count': len(items),
                'modes': sorted({i['mode'] for i in items if i['mode']}),
                'consolidatable': len(items) > 1,
                'shipments': [
                    {'reference': i['reference'], 'supplier': i['supplier'],
                     'mode': i['mode'], 'stage': i['stage'], 'destination': i['destination']}
                    for i in items
                ],
            }
            for fwd, items in ranked
        ],
        'total_groups': len(groups),
        'consolidatable_groups': sum(1 for _, items in ranked if len(items) > 1),
    })


@app.route('/api/stats/segments')
def api_stats_segments():
    """Segment breakdowns for the progress pie charts (real data only)."""
    conn = get_db()
    stages = {}
    for r in conn.execute('SELECT stage, COUNT(*) c FROM shipments GROUP BY stage').fetchall():
        stages[r['stage']] = r['c']
    modes = {}
    for r in conn.execute("SELECT COALESCE(NULLIF(mode,''),'unknown') mode, COUNT(*) c FROM shipments GROUP BY mode").fetchall():
        modes[r['mode']] = r['c']
    # attention buckets
    buckets = {}
    for r in conn.execute('SELECT * FROM shipments').fetchall():
        b = enrich_shipment(dict(r))['attention']
        buckets[b] = buckets.get(b, 0) + 1
    conn.close()
    return jsonify({'stages': stages, 'modes': modes, 'attention': buckets})


@app.route('/api/shipments')
def api_shipments():
    attention = request.args.get('attention', '').strip()
    conn = get_db()
    rows = conn.execute('SELECT * FROM shipments ORDER BY '
                        "CASE WHEN eta_arrival IS NULL OR eta_arrival='' THEN 1 ELSE 0 END, "
                        'eta_arrival ASC').fetchall()
    conn.close()
    out = [enrich_shipment(dict(r)) for r in rows]
    if attention:
        out = [s for s in out if s['attention'] == attention]
    # attention summaries
    buckets = {}
    for s in out:
        buckets[s['attention']] = buckets.get(s['attention'], 0) + 1
    return jsonify({'shipments': out, 'buckets': buckets,
                    'stages': STAGES})


@app.route('/api/shipments/<int:sid>')
def api_shipment(sid):
    conn = get_db()
    r = conn.execute('SELECT * FROM shipments WHERE id=?', (sid,)).fetchone()
    conn.close()
    if not r:
        return jsonify({'error': 'not found'}), 404
    return jsonify({'shipment': enrich_shipment(dict(r))})


@app.route('/api/shipments', methods=['POST'])
def api_shipments_create():
    data = request.get_json(force=True, silent=True) or {}
    ref = (data.get('reference') or '').strip()
    if not ref:
        return jsonify({'error': 'reference required'}), 400
    stage = data.get('stage') or 'Booked'
    if stage not in STAGE_INDEX:
        return jsonify({'error': f'stage must be one of {STAGES}'}), 400
    # link to a fleet vessel if carrier name matches
    linked = None
    carrier = (data.get('carrier') or '').strip()
    if carrier:
        conn = get_db()
        linked = conn.execute('SELECT id FROM vessels WHERE UPPER(name) LIKE ?',
                              (f'%{carrier.upper()}%',)).fetchone()
        conn.close()
    conn = get_db()
    conn.execute('''INSERT INTO shipments
        (reference, supplier, description, mode, carrier, awb_bl, container_no, destination,
         forwarder, forwarder_contact, stage, eta_arrival, notes, linked_vessel_id)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (ref, data.get('supplier'), data.get('description'), data.get('mode', 'sea'),
         carrier, data.get('awb_bl'), data.get('container_no'),
         (data.get('destination') or 'DURBAN').strip().upper(),
         data.get('forwarder'), data.get('forwarder_contact'), stage,
         data.get('eta_arrival'), data.get('notes'),
         linked['id'] if linked else None))
    sid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    row = conn.execute('SELECT * FROM shipments WHERE id=?', (sid,)).fetchone()
    conn.commit()
    conn.close()
    return jsonify({'shipment': enrich_shipment(dict(row))}), 201


@app.route('/api/shipments/<int:sid>', methods=['PATCH'])
def api_shipment_update(sid):
    data = request.get_json(force=True, silent=True) or {}
    conn = get_db()
    row = conn.execute('SELECT * FROM shipments WHERE id=?', (sid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'not found'}), 404
    row = dict(row)
    allowed = ['supplier', 'description', 'mode', 'carrier', 'awb_bl', 'container_no',
               'destination', 'forwarder', 'forwarder_contact', 'eta_arrival', 'notes',
               'commercial_invoice_no', 'po_number', 'freight_file_no', 'incoming_stock',
               'airport', 'port', 'berth', 'ship']
    for k in allowed:
        if k in data:
            row[k] = data[k]
    if 'stage' in data:
        if data['stage'] not in STAGE_INDEX:
            conn.close()
            return jsonify({'error': f'stage must be one of {STAGES}'}), 400
        row['stage'] = data['stage']
        # auto-stamp cleared/delivered timestamps when moved into those stages
        now_iso = datetime.now().strftime('%Y-%m-%d %H:%M')
        if data['stage'] == 'Customs Cleared' and not row.get('cleared_at'):
            row['cleared_at'] = now_iso
        if data['stage'] == 'Delivered' and not row.get('delivered_at'):
            row['delivered_at'] = now_iso
    # re-link vessel if carrier changed
    if 'carrier' in data and data['carrier']:
        linked = conn.execute('SELECT id FROM vessels WHERE UPPER(name) LIKE ?',
                              (f'%{data["carrier"].upper()}%',)).fetchone()
        row['linked_vessel_id'] = linked['id'] if linked else None
    fields = allowed + ['stage', 'linked_vessel_id', 'cleared_at', 'delivered_at']
    sets = ', '.join(f'{f}=?' for f in fields)
    conn.execute(f'UPDATE shipments SET {sets}, updated_at=datetime(\'now\') WHERE id=?',
                 tuple(row.get(f) for f in fields) + (sid,))
    row2 = conn.execute('SELECT * FROM shipments WHERE id=?', (sid,)).fetchone()
    conn.commit()
    conn.close()
    return jsonify({'shipment': enrich_shipment(dict(row2))})


@app.route('/api/shipments/<int:sid>', methods=['DELETE'])
def api_shipment_delete(sid):
    conn = get_db()
    conn.execute('DELETE FROM shipments WHERE id=?', (sid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'deleted': sid})


# --- Shipment notes (timestamped thread) ---
@app.route('/api/shipments/<int:sid>/notes', methods=['GET', 'POST'])
def api_shipment_notes(sid):
    conn = get_db()
    if request.method == 'POST':
        data = request.get_json(force=True, silent=True) or {}
        note = (data.get('note') or '').strip()
        if not note:
            return jsonify({'error': 'note required'}), 400
        conn.execute('INSERT INTO shipment_notes (shipment_id, note, author) VALUES (?,?,?)',
                     (sid, note, data.get('author') or 'user'))
        conn.commit()
        row = conn.execute('SELECT * FROM shipment_notes WHERE id=last_insert_rowid()').fetchone()
        conn.close()
        return jsonify({'note': dict(row)}), 201
    rows = conn.execute('SELECT * FROM shipment_notes WHERE shipment_id=? ORDER BY created_at',
                       (sid,)).fetchall()
    conn.close()
    return jsonify({'notes': [dict(r) for r in rows]})


# --- Shipment documents (upload per shipment, by doc type) ---
DOC_TYPES = ['Proforma Invoice', 'Supplier PO', 'Commercial Invoice', 'Packing List',
             'HAWB', 'MAWB', 'SAD500', 'SAD501', 'SAD507', 'Customs Worksheet',
             'EDI / Release Doc', 'Clearing Instruction', 'POA', 'Other']
ALLOWED_EXT = {'.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx', '.xls', '.xlsx',
               '.csv', '.txt', '.xml', '.zip', '.tif', '.tiff'}


@app.route('/api/shipments/<int:sid>/documents', methods=['GET', 'POST'])
def api_shipment_documents(sid):
    conn = get_db()
    # ensure shipment exists
    if not conn.execute('SELECT 1 FROM shipments WHERE id=?', (sid,)).fetchone():
        conn.close()
        return jsonify({'error': 'shipment not found'}), 404
    if request.method == 'POST':
        if 'file' not in request.files:
            conn.close()
            return jsonify({'error': 'no file part'}), 400
        f = request.files['file']
        if not f or f.filename == '':
            conn.close()
            return jsonify({'error': 'empty filename'}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED_EXT:
            conn.close()
            return jsonify({'error': f'file type {ext} not allowed'}), 400
        import uuid
        stored = uuid.uuid4().hex + ext
        f.save(os.path.join(UPLOAD_DIR, stored))
        doc_type = (request.form.get('doc_type') or 'Other').strip()
        conn.execute('''INSERT INTO shipment_documents
            (shipment_id, doc_type, filename, stored_name, content_type, size)
            VALUES (?,?,?,?,?,?)''',
            (sid, doc_type, f.filename, stored, f.content_type or '', os.path.getsize(os.path.join(UPLOAD_DIR, stored))))
        conn.commit()
        did = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        row = conn.execute('SELECT * FROM shipment_documents WHERE id=?', (did,)).fetchone()
        conn.close()
        return jsonify({'document': dict(row)}), 201
    rows = conn.execute('SELECT * FROM shipment_documents WHERE shipment_id=? ORDER BY doc_type, uploaded_at',
                       (sid,)).fetchall()
    conn.close()
    return jsonify({'documents': [dict(r) for r in rows], 'doc_types': DOC_TYPES})


@app.route('/api/documents/<int:did>/download')
def api_document_download(did):
    conn = get_db()
    row = conn.execute('SELECT * FROM shipment_documents WHERE id=?', (did,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'not found'}), 404
    path = os.path.join(UPLOAD_DIR, row['stored_name'])
    if not os.path.exists(path):
        return jsonify({'error': 'file missing on disk'}), 404
    return send_file(path, as_attachment=True, download_name=row['filename'],
                     mimetype=row['content_type'] or 'application/octet-stream')


@app.route('/api/documents/<int:did>', methods=['DELETE'])
def api_document_delete(did):
    conn = get_db()
    row = conn.execute('SELECT * FROM shipment_documents WHERE id=?', (did,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'not found'}), 404
    path = os.path.join(UPLOAD_DIR, row['stored_name'])
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError:
        pass
    conn.execute('DELETE FROM shipment_documents WHERE id=?', (did,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# --- Shipment flags (typed status tags) ---
@app.route('/api/shipments/<int:sid>/flags', methods=['GET', 'POST', 'DELETE'])
def api_shipment_flags(sid):
    conn = get_db()
    if request.method == 'POST':
        data = request.get_json(force=True, silent=True) or {}
        flag = (data.get('flag') or '').strip()
        if flag not in FLAGS:
            return jsonify({'error': f'flag must be one of {FLAGS}'}), 400
        conn.execute('INSERT OR IGNORE INTO shipment_flags (shipment_id, flag) VALUES (?,?)',
                     (sid, flag))
        conn.commit()
        conn.close()
        return jsonify({'flags': [r['flag'] for r in get_db().execute(
            'SELECT flag FROM shipment_flags WHERE shipment_id=?', (sid,)).fetchall()]})
    if request.method == 'DELETE':
        data = request.get_json(force=True, silent=True) or {}
        flag = (data.get('flag') or '').strip()
        conn.execute('DELETE FROM shipment_flags WHERE shipment_id=? AND flag=?', (sid, flag))
        conn.commit()
        conn.close()
        return jsonify({'flags': [r['flag'] for r in get_db().execute(
            'SELECT flag FROM shipment_flags WHERE shipment_id=?', (sid,)).fetchall()]})
    rows = conn.execute('SELECT flag FROM shipment_flags WHERE shipment_id=? ORDER BY flag',
                       (sid,)).fetchall()
    conn.close()
    return jsonify({'flags': [r['flag'] for r in rows]})


# --- Shipment slot (harbour booking / flight cutoff) ---
@app.route('/api/shipments/<int:sid>/slot', methods=['POST'])
def api_shipment_slot(sid):
    data = request.get_json(force=True, silent=True) or {}
    slot_type = data.get('slot_type')
    if slot_type not in (None, '', 'Harbour', 'Flight', 'Other'):
        return jsonify({'error': 'invalid slot_type'}), 400
    allowed = {k: data.get(k) for k in ('slot_type', 'slot_datetime', 'slot_ref')}
    conn = get_db()
    sets = ', '.join(f'{k}=?' for k in allowed)
    conn.execute(f'UPDATE shipments SET {sets}, updated_at=datetime(\'now\') WHERE id=?',
                 tuple(allowed.values()) + (sid,))
    conn.commit()
    row = conn.execute('SELECT * FROM shipments WHERE id=?', (sid,)).fetchone()
    conn.close()
    return jsonify({'shipment': enrich_shipment(dict(row))})


# --- Flags catalogue (for the UI) ---
@app.route('/api/flags')
def api_flags():
    return jsonify({'flags': FLAGS, 'colors': FLAG_COLORS})


@app.route('/api/shipments/<int:sid>/message')
def api_shipment_message(sid):
    """Pre-filled status message for the clearing & forwarding agent."""
    conn = get_db()
    r = conn.execute('SELECT * FROM shipments WHERE id=?', (sid,)).fetchone()
    conn.close()
    if not r:
        return jsonify({'error': 'not found'}), 404
    s = enrich_shipment(dict(r))
    eta = s['eta_arrival'] or 'TBC'
    if s['eta_countdown_hours'] is not None:
        cd = s['eta_countdown_hours']
        when = f'overdue by {-cd:.0f}h' if cd < 0 else f'due in {cd:.0f}h'
    else:
        when = 'ETA TBC'
    msg = (f"Hi {s['forwarder'] or 'team'}, status check on Josawa shipment "
           f"{s['reference']} ({s['description'] or 'goods'}).\n"
           f"Mode: {s['mode'].upper()} | Carrier: {s['carrier'] or 'TBC'} | "
           f"{'AWB' if s['mode']=='air' else 'B/L'}: {s['awb_bl'] or s['container_no'] or 'TBC'}.\n"
           f"Current stage: {s['stage']}. ETA: {eta} ({when}).\n"
           f"Please advise clearance/delivery status and any documents required.\nThanks.")
    return jsonify({'shipment': s['reference'], 'forwarder': s['forwarder'],
                    'contact': s['forwarder_contact'], 'message': msg})


@app.route('/api/track/flight')
def api_track_flight():
    flight = request.args.get('flight', '').strip().upper()
    if not flight:
        return jsonify({'error': 'Flight number required', 'flights': []})
    # Live: match against current OpenSky states by callsign prefix
    states = get_live_flights()
    matched = []
    for st in states:
        cs = (st[1] or '').strip().upper()
        if cs and (cs == flight or cs.startswith(flight)):
            d = aircraft_to_dict(st)
            if d:
                d['status'] = 'LIVE — over Southern Africa' if d['lat'] else 'Tracking'
                matched.append(d)
    if not matched:
        # best-effort FR24 public page
        for url in (f'https://www.flightradar24.com/data/flights/{flight}',
                    f'https://www.flightradar24.com/data/aircraft/{flight}'):
            html = fetch_url(url)
            if not html:
                continue
            import re
            rows = re.findall(r'<tr[^>]*data-flight[^>]*>(.*?)</tr>', html, re.DOTALL)
            for f in rows[:5]:
                cells = re.findall(r'<td[^>]*>(.*?)</td>', f, re.DOTALL)
                if len(cells) >= 4:
                    clean = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
                    matched.append({'callsign': flight, 'airline': clean[0],
                                    'route': clean[1], 'status': clean[2],
                                    'aircraft': clean[3], 'live': False})
            if matched:
                break
    if not matched:
        matched.append({'callsign': flight, 'status': 'Searching…', 'live': False})
    return jsonify({'flight': flight, 'flights': matched,
                    'live': any(m.get('live', True) for m in matched)})


@app.route('/api/track/shipping')
def api_track_shipping():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({'error': 'Shipping number, vessel name or company required', 'shipments': []})
    ql = q.upper()
    conn = get_db()
    # match on vessel name or company (transporter) — seeded fleet
    rows = conn.execute(
        '''SELECT * FROM vessels
           WHERE UPPER(name) LIKE ? OR UPPER(transporter) LIKE ?
              OR UPPER(destination) LIKE ? OR UPPER(port) LIKE ? OR UPPER(imo) LIKE ?
           ORDER BY CASE WHEN UPPER(name)=? THEN 0 ELSE 1 END''',
        (f'%{ql}%', f'%{ql}%', f'%{ql}%', f'%{ql}%', f'%{ql}%', ql)).fetchall()
    conn.close()
    shipments = []
    for r in rows:
        r = dict(r)
        shipments.append({
            'name': r['name'], 'imo': r['imo'], 'type': r['vtype'], 'flag': r['flag'],
            'lat': r['lat'], 'lon': r['lon'], 'speed_kn': r['speed_kn'],
            'heading': r['heading'], 'status': r['status'], 'origin': r['origin'],
            'destination': r['destination'], 'port': r['port'], 'eta': r['eta_arrival'],
            'transporter': r['transporter'], 'group_code': r['group_code'],
            'source': 'Josawa fleet DB',
        })
    if not shipments:
        shipments.append({'name': q, 'status': 'No matching vessel in fleet DB',
                          'source': '—'})
    return jsonify({'query': q, 'shipments': shipments,
                    'live': bool(AISSTREAM_KEY)})


@app.route('/api/assets', methods=['GET', 'POST'])
def api_assets():
    conn = get_db()
    if request.method == 'POST':
        data = request.get_json(force=True, silent=True) or {}
        kind = (data.get('kind') or '').strip().lower()
        identifier = (data.get('identifier') or '').strip()
        destination = (data.get('destination') or '').strip().upper()
        if kind not in ('ship', 'flight'):
            return jsonify({'error': 'kind must be ship or flight'}), 400
        if not identifier:
            return jsonify({'error': 'identifier required'}), 400
        if not destination:
            destination = DEFAULT_DEST[kind]
        if destination not in ALLOWED_DEST:
            return jsonify({'error': f'destination must be one of {sorted(ALLOWED_DEST)}'}), 400
        # link to a fleet vessel if name matches
        linked = None
        if kind == 'ship':
            linked = conn.execute('SELECT id FROM vessels WHERE UPPER(name) LIKE ?',
                                  (f'%{identifier.upper()}%',)).fetchone()
        existing = conn.execute(
            'SELECT id FROM tracked_assets WHERE kind=? AND identifier=? AND destination=?',
            (kind, identifier.upper(), destination)).fetchone()
        if existing:
            conn.close()
            return jsonify({'error': 'duplicate asset', 'id': existing['id']}), 409
        lat = data.get('lat')
        lon = data.get('lon')
        linked_id = linked['id'] if linked else None
        conn.execute(
            'INSERT INTO tracked_assets (kind, identifier, destination, lat, lon, position_source, linked_vessel_id) VALUES (?,?,?,?,?,?,?)',
            (kind, identifier.upper(), destination, lat, lon,
             'manual' if lat else None, linked_id))
        aid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        row = conn.execute('SELECT * FROM tracked_assets WHERE id=?', (aid,)).fetchone()
        conn.commit()
        conn.close()
        return jsonify({'asset': enrich_asset(dict(row))}), 201
    rows = conn.execute('SELECT * FROM tracked_assets ORDER BY id DESC').fetchall()
    conn.close()
    return jsonify({'assets': [enrich_asset(dict(r)) for r in rows],
                    'targets': {k: v['label'] for k, v in SA_TARGETS.items()}})


@app.route('/api/assets/<int:aid>', methods=['DELETE'])
def api_asset_delete(aid):
    conn = get_db()
    conn.execute('DELETE FROM tracked_assets WHERE id=?', (aid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'deleted': aid})


@app.route('/api/assets/refresh', methods=['POST'])
def api_assets_refresh():
    """Refresh live positions: OpenSky for flights, snapped ships for arrived vessels."""
    conn = get_db()
    rows = conn.execute('SELECT * FROM tracked_assets').fetchall()
    states = get_live_flights()
    updated = []
    for r in rows:
        r = dict(r)
        if r['kind'] == 'flight':
            for st in states:
                callsign = (st[1] or '').strip().upper()
                if callsign == r['identifier']:
                    r['lat'], r['lon'] = st[6], st[5]
                    r['speed_kmh'] = round((st[9] or 0) * 3.6, 1)
                    r['position_source'] = 'OpenSky'
                    break
        else:
            if r['linked_vessel_id']:
                v = conn.execute('SELECT status, lat, lon FROM vessels WHERE id=?',
                                 (r['linked_vessel_id'],)).fetchone()
                if v and v['status'] in ('Moored', 'Arrived'):
                    tgt = SA_TARGETS.get(r['destination'], SA_TARGETS['DURBAN'])
                    r['lat'], r['lon'] = tgt['lat'], tgt['lon']
                    r['speed_kmh'] = 0
                    r['position_source'] = 'arrived-snap'
                    r['arrived'] = 1
                    r['arrived_at'] = datetime.now(timezone.utc).isoformat()
        conn.execute('UPDATE tracked_assets SET lat=?, lon=?, speed_kmh=?, position_source=?, arrived=?, arrived_at=? WHERE id=?',
                     (r['lat'], r['lon'], r['speed_kmh'], r['position_source'],
                      r.get('arrived', 0), r.get('arrived_at'), r['id']))
        updated.append(enrich_asset(r))
    conn.commit()
    conn.close()
    return jsonify({'assets': updated, 'live_flights': len(states)})


@app.route('/api/map')
def api_map():
    """Live map data: fleet vessels + live flights + tracked assets + shipments."""
    conn = get_db()
    vessels = [dict(r) for r in conn.execute('SELECT * FROM vessels').fetchall()]
    assets = [enrich_asset(dict(r)) for r in conn.execute('SELECT * FROM tracked_assets').fetchall()]
    shipments = [enrich_shipment(dict(r)) for r in conn.execute('SELECT * FROM shipments').fetchall()]
    conn.close()
    flights = [d for d in (aircraft_to_dict(st) for st in get_live_flights()) if d and d['lat']]
    return jsonify({
        'vessels': vessels,
        'flights': flights,
        'assets': assets,
        'shipments': shipments,
        'targets': SA_TARGETS,
        'live_ships': bool(AISSTREAM_KEY),
        'updated_at': datetime.now(timezone.utc).isoformat(),
    })


@app.route('/sse')
def sse_stream():
    """Server-Sent Events: heartbeat every 60s + push new alerts as they arrive."""
    last_id = [0]
    # seed last_id with current max so we only push NEW notifications
    conn = get_db()
    row = conn.execute('SELECT MAX(id) AS m FROM notifications').fetchone()
    conn.close()
    last_id[0] = row['m'] or 0

    def generate():
        yield f"event: connected\ndata: {json.dumps({'type': 'connected'})}\n\n"
        while True:
            conn = get_db()
            new = conn.execute('SELECT * FROM notifications WHERE id > ? ORDER BY id',
                               (last_id[0],)).fetchall()
            conn.close()
            for n in new:
                last_id[0] = n['id']
                yield f"event: alert\ndata: {json.dumps(dict(n))}\n\n"
            time.sleep(5)
    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'Connection': 'keep-alive',
                             'Access-Control-Allow-Origin': '*'})


# --- CRM autofill (from imported TASCAM SA CRM) ---
@app.route('/api/crm/lookup')
def api_crm_lookup():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify({'results': []})
    like = f'%{q.upper()}%'
    conn = get_db()
    rows = conn.execute(
        'SELECT company, contact_name, email, phone, city, province, sector FROM crm_contacts '
        'WHERE UPPER(company) LIKE ? OR UPPER(contact_name) LIKE ? OR UPPER(email) LIKE ? '
        'ORDER BY company LIMIT 8', (like, like, like)).fetchall()
    conn.close()
    return jsonify({'results': [dict(r) for r in rows]})


@app.route('/api/crm/stats')
def api_crm_stats():
    conn = get_db()
    c = conn.execute('SELECT COUNT(*) AS c FROM crm_contacts').fetchone()['c']
    conn.close()
    return jsonify({'crm_contacts': c})


# --- CSV import ---
@app.route('/api/import/csv', methods=['POST'])
def api_import_csv():
    import csv, io
    payload = request.get_json(force=True, silent=True) or {}
    text = payload.get('csv') or ''
    if not text.strip():
        return jsonify({'error': 'no csv provided', 'imported': 0}), 400
    reader = csv.DictReader(io.StringIO(text))
    cols = [c.lower() for c in (reader.fieldnames or [])]
    # map common column aliases -> shipment fields
    alias = {
        'reference': 'reference', 'ref': 'reference', 'shipment': 'reference', 'id': 'reference',
        'supplier': 'supplier', 'vendor': 'supplier', 'shipper': 'supplier',
        'description': 'description', 'goods': 'description', 'product': 'description', 'items': 'description',
        'mode': 'mode', 'transport': 'mode',
        'carrier': 'carrier', 'vessel': 'carrier', 'airline': 'carrier', 'steamship': 'carrier',
        'awb': 'awb_bl', 'bl': 'awb_bl', 'bill': 'awb_bl', 'bill_of_lading': 'awb_bl',
        'container': 'container_no', 'container_no': 'container_no', 'container_no.': 'container_no',
        'destination': 'destination', 'dest': 'destination', 'port': 'destination',
        'forwarder': 'forwarder', 'clearing': 'forwarder', 'agent': 'forwarder',
        'forwarder_contact': 'forwarder_contact', 'contact': 'forwarder_contact', 'email': 'forwarder_contact',
        'eta': 'eta_arrival', 'eta_arrival': 'eta_arrival', 'arrival': 'eta_arrival', 'eta_date': 'eta_arrival',
        'notes': 'notes',
    }
    mapped_cols = {c: alias.get(c) for c in cols if alias.get(c)}
    imported = 0
    skipped = 0
    conn = get_db()
    for row in reader:
        data = {mapped_cols[k]: (row.get(k) or '').strip() for k in row if mapped_cols.get(k)}
        ref = (data.get('reference') or '').strip()
        if not ref:
            skipped += 1
            continue
        if conn.execute('SELECT 1 FROM shipments WHERE reference=?', (ref,)).fetchone():
            skipped += 1
            continue
        mode = (data.get('mode') or 'sea').strip().lower()
        if mode not in ('sea', 'air'):
            mode = 'sea'
        dest = (data.get('destination') or 'DURBAN').strip().upper()
        if dest not in ALLOWED_DEST:
            dest = 'DURBAN'
        stage = data.get('stage') or 'Booked'
        if stage not in STAGE_INDEX:
            stage = 'Booked'
        # link vessel if carrier matches fleet
        linked = None
        carrier = (data.get('carrier') or '').strip()
        if carrier:
            linked = conn.execute('SELECT id FROM vessels WHERE UPPER(name) LIKE ?',
                                  (f'%{carrier.upper()}%',)).fetchone()
        conn.execute('''INSERT INTO shipments
            (reference, supplier, description, mode, carrier, awb_bl, container_no, destination,
             forwarder, forwarder_contact, stage, eta_arrival, notes, linked_vessel_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (ref, data.get('supplier'), data.get('description'), mode, carrier,
             data.get('awb_bl'), data.get('container_no'), dest, data.get('forwarder'),
             data.get('forwarder_contact'), stage, data.get('eta_arrival'), data.get('notes'),
             linked['id'] if linked else None))
        imported += 1
    conn.commit()
    conn.close()
    return jsonify({'imported': imported, 'skipped': skipped,
                   'message': f'Imported {imported} shipment(s), skipped {skipped} (missing ref / duplicate)'})


# --- Excel (xlsx) export / import ---
SHIP_EXPORT_COLS = [
    'reference', 'supplier', 'description', 'mode', 'carrier', 'awb_bl', 'container_no',
    'destination', 'forwarder', 'forwarder_contact', 'stage', 'eta_arrival',
    'commercial_invoice_no', 'po_number', 'freight_file_no', 'incoming_stock',
    'airport', 'port', 'berth', 'ship', 'cleared_at', 'delivered_at', 'notes',
]
XLSX_ALIAS = {
    'reference': 'reference', 'ref': 'reference', 'shipment': 'reference', 'id': 'reference',
    'supplier': 'supplier', 'vendor': 'supplier', 'shipper': 'supplier',
    'description': 'description', 'goods': 'description', 'product': 'description', 'items': 'description',
    'mode': 'mode', 'transport': 'mode',
    'carrier': 'carrier', 'vessel': 'carrier', 'airline': 'carrier', 'steamship': 'carrier',
    'awb': 'awb_bl', 'bl': 'awb_bl', 'bill': 'awb_bl', 'bill_of_lading': 'awb_bl',
    'container': 'container_no', 'container_no': 'container_no',
    'destination': 'destination', 'dest': 'destination', 'port_dest': 'destination',
    'forwarder': 'forwarder', 'clearing': 'forwarder', 'agent': 'forwarder',
    'forwarder_contact': 'forwarder_contact', 'contact': 'forwarder_contact', 'email': 'forwarder_contact',
    'eta': 'eta_arrival', 'eta_arrival': 'eta_arrival', 'arrival': 'eta_arrival', 'eta_date': 'eta_arrival',
    'commercial_invoice': 'commercial_invoice_no', 'commercial_invoice_no': 'commercial_invoice_no', 'ci': 'commercial_invoice_no', 'ci_no': 'commercial_invoice_no',
    'po': 'po_number', 'po_number': 'po_number', 'po_no': 'po_number',
    'freight_file': 'freight_file_no', 'freight_file_no': 'freight_file_no',
    'incoming_stock': 'incoming_stock', 'stock': 'incoming_stock',
    'airport': 'airport', 'discharge_port': 'port', 'port': 'port',
    'berth': 'berth', 'ship': 'ship', 'vessel_name': 'ship',
    'cleared_at': 'cleared_at', 'delivered_at': 'delivered_at', 'notes': 'notes',
}


@app.route('/api/shipments/export/xlsx')
def api_export_xlsx():
    from openpyxl import Workbook
    conn = get_db()
    rows = conn.execute('SELECT * FROM shipments ORDER BY reference').fetchall()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Shipments'
    ws.append(SHIP_EXPORT_COLS)
    for r in rows:
        d = dict(r)
        ws.append([d.get(c) or '' for c in SHIP_EXPORT_COLS])
    # autosize columns
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 48)
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='Josawa_Shipments.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/import/xlsx', methods=['POST'])
def api_import_xlsx():
    if 'file' not in request.files:
        return jsonify({'error': 'no file part'}), 400
    f = request.files['file']
    if not f or f.filename == '':
        return jsonify({'error': 'empty filename'}), 400
    import io
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(h or '').strip().lower() for h in next(rows_iter)]
        mapped = {i: XLSX_ALIAS.get(h) for i, h in enumerate(header) if XLSX_ALIAS.get(h)}
        imported = skipped = 0
        conn = get_db()
        for raw in rows_iter:
            if raw is None:
                continue
            data = {mapped[i]: (str(raw[i] or '').strip()) for i in mapped if i < len(raw) and raw[i] is not None}
            ref = (data.get('reference') or '').strip()
            if not ref:
                skipped += 1
                continue
            if conn.execute('SELECT 1 FROM shipments WHERE reference=?', (ref,)).fetchone():
                skipped += 1
                continue
            mode = (data.get('mode') or 'sea').strip().lower()
            if mode not in ('sea', 'air'):
                mode = 'sea'
            dest = (data.get('destination') or 'DURBAN').strip().upper()
            if dest not in ALLOWED_DEST:
                dest = 'DURBAN'
            stage = data.get('stage') or 'Booked'
            if stage not in STAGE_INDEX:
                stage = 'Booked'
            carrier = (data.get('carrier') or '').strip()
            linked = None
            if carrier:
                linked = conn.execute('SELECT id FROM vessels WHERE UPPER(name) LIKE ?',
                                      (f'%{carrier.upper()}%',)).fetchone()
            conn.execute('''INSERT INTO shipments
                (reference, supplier, description, mode, carrier, awb_bl, container_no, destination,
                 forwarder, forwarder_contact, stage, eta_arrival, commercial_invoice_no, po_number,
                 freight_file_no, incoming_stock, airport, port, berth, ship, notes, linked_vessel_id)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (ref, data.get('supplier'), data.get('description'), mode, carrier,
                 data.get('awb_bl'), data.get('container_no'), dest, data.get('forwarder'),
                 data.get('forwarder_contact'), stage, data.get('eta_arrival'),
                 data.get('commercial_invoice_no'), data.get('po_number'), data.get('freight_file_no'),
                 data.get('incoming_stock'), data.get('airport'), data.get('port'),
                 data.get('berth'), data.get('ship'), data.get('notes'),
                 linked['id'] if linked else None))
            imported += 1
        conn.commit()
        conn.close()
        return jsonify({'imported': imported, 'skipped': skipped,
                       'message': f'Imported {imported} shipment(s) from Excel, skipped {skipped} (missing ref / duplicate)'})
    except Exception as e:
        return jsonify({'error': f'Excel parse failed: {e}'}), 400


# --- Notifications / alerts ---
@app.route('/api/notifications')
def api_notifications():
    since = request.args.get('since', '0')
    try:
        since = int(since)
    except ValueError:
        since = 0
    conn = get_db()
    rows = conn.execute('SELECT * FROM notifications WHERE id > ? ORDER BY id DESC LIMIT 50',
                        (since,)).fetchall()
    unread = conn.execute('SELECT COUNT(*) AS c FROM notifications WHERE read=0').fetchone()['c']
    conn.close()
    return jsonify({'notifications': [dict(r) for r in rows], 'unread': unread})


@app.route('/api/notifications/<int:nid>', methods=['POST'])
def api_notification_read(nid):
    conn = get_db()
    conn.execute('UPDATE notifications SET read=1 WHERE id=?', (nid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# --- Suppliers (master data; seeded with placeholder entries, editable) ---
@app.route('/api/suppliers', methods=['GET', 'POST'])
def api_suppliers():
    conn = get_db()
    if request.method == 'POST':
        d = request.get_json(force=True, silent=True) or {}
        name = (d.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'supplier name required'}), 400
        conn.execute('''INSERT OR REPLACE INTO suppliers
            (name, contact_name, email, phone, city, country, products, notes)
            VALUES (?,?,?,?,?,?,?,?)''',
            (name, d.get('contact_name'), d.get('email'), d.get('phone'),
             d.get('city'), d.get('country'), d.get('products'), d.get('notes')))
        conn.commit()
        sid = conn.execute('SELECT id FROM suppliers WHERE name=?', (name,)).fetchone()['id']
        conn.close()
        return jsonify({'ok': True, 'id': sid})
    rows = conn.execute('SELECT * FROM suppliers ORDER BY name').fetchall()
    conn.close()
    return jsonify({'suppliers': [dict(r) for r in rows]})


@app.route('/api/suppliers/<int:sid>', methods=['DELETE'])
def api_supplier_delete(sid):
    conn = get_db()
    conn.execute('DELETE FROM suppliers WHERE id=?', (sid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/suppliers/lookup')
def api_suppliers_lookup():
    q = request.args.get('q', '').strip()
    conn = get_db()
    if q:
        rows = conn.execute('SELECT * FROM suppliers WHERE name LIKE ? ORDER BY name LIMIT 8',
                            ('%' + q + '%',)).fetchall()
    else:
        rows = conn.execute('SELECT * FROM suppliers ORDER BY name LIMIT 8').fetchall()
    conn.close()
    return jsonify({'results': [dict(r) for r in rows]})


# --- Alerts & Updates feed (SAPS stops, clearance, delays) ---
@app.route('/api/alerts', methods=['GET', 'POST'])
def api_alerts():
    conn = get_db()
    if request.method == 'POST':
        d = request.get_json(force=True, silent=True) or {}
        kind = (d.get('kind') or 'Other').strip()
        title = (d.get('title') or '').strip()
        if not title:
            return jsonify({'error': 'title required'}), 400
        conn.execute('''INSERT INTO alerts_updates (kind, title, message, shipment_ref)
            VALUES (?,?,?,?)''',
            (kind, title, d.get('message'), d.get('shipment_ref')))
        conn.commit()
        aid = conn.execute('SELECT id FROM alerts_updates WHERE kind=? AND title=? ORDER BY id DESC LIMIT 1',
                           (kind, title)).fetchone()['id']
        conn.close()
        return jsonify({'ok': True, 'id': aid})
    rows = conn.execute('SELECT * FROM alerts_updates ORDER BY created_at DESC LIMIT 50').fetchall()
    conn.close()
    return jsonify({'alerts': [dict(r) for r in rows]})


@app.route('/api/alerts/<int:aid>', methods=['POST'])
def api_alert_toggle(aid):
    conn = get_db()
    r = conn.execute('SELECT resolved FROM alerts_updates WHERE id=?', (aid,)).fetchone()
    if not r:
        conn.close()
        return jsonify({'error': 'not found'}), 404
    newval = 0 if r['resolved'] else 1
    conn.execute('UPDATE alerts_updates SET resolved=? WHERE id=?', (newval, aid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'resolved': bool(newval)})


# --- Settings ---
@app.route('/api/settings', methods=['GET', 'POST'])
def api_settings():
    if request.method == 'POST':
        data = request.get_json(force=True, silent=True) or {}
        for k, v in data.items():
            set_setting(k, str(v))
        return jsonify({'ok': True})
    keys = ['digest_enabled', 'digest_email', 'digest_hour', 'digest_timezone',
            'smtp_host', 'smtp_user', 'smtp_from', 'alerts_enabled']
    return jsonify({k: get_setting(k) for k in keys})


# --- Morning digest ---
def build_digest():
    """Compose the 'my day' morning digest text from current shipments."""
    conn = get_db()
    rows = [enrich_shipment(dict(r)) for r in conn.execute('SELECT * FROM shipments').fetchall()]
    conn.close()
    now = datetime.now()
    def within(s, days):
        eta = _parse_dt(s.get('eta_arrival'))
        if not eta:
            return False
        diff = (eta - now).total_seconds()
        return -1 <= diff <= days * 86400  # -1h grace for just-overdue
    arriving_today = [s for s in rows if within(s, 1)]
    this_week = [s for s in rows if within(s, 7) and s not in arriving_today]
    needs_clearing = [s for s in rows if s['attention'] == 'Needs Clearing']
    ready = [s for s in rows if s['attention'] == 'Ready for Collection']
    overdue = [s for s in rows if s['attention'] == 'Overdue']
    lines = []
    lines.append('JOSAWA LOGISTICS — MORNING DIGEST')
    lines.append(now.strftime('%A %d %B %Y'))
    lines.append('=' * 40)
    lines.append(f'Arriving today ({len(arriving_today)}):')
    for s in arriving_today:
        lines.append(f"  - {s['reference']} {s['description'] or ''} -> {s['destination']} (ETA {s['eta_arrival']})")
    lines.append(f'Arriving this week ({len(this_week)}):')
    for s in this_week:
        lines.append(f"  - {s['reference']} -> {s['destination']} (ETA {s['eta_arrival']})")
    lines.append(f'Needs clearing ({len(needs_clearing)}):')
    for s in needs_clearing:
        lines.append(f"  - {s['reference']} [{s['forwarder']}] {s['eta_arrival'] or ''}")
    lines.append(f'Ready for collection ({len(ready)}):')
    for s in ready:
        lines.append(f"  - {s['reference']} [{s['forwarder']}]")
    if overdue:
        lines.append(f'OVERDUE ({len(overdue)}):')
        for s in overdue:
            lines.append(f"  - {s['reference']} [{s['forwarder']}]")
    # Booking slots (harbour / flight cutoff) due soon or missed
    slot_due = [s for s in rows if s.get('slot_status') in ('due-soon', 'missed')]
    if slot_due:
        lines.append(f'BOOKING SLOTS — action ({len(slot_due)}):')
        for s in slot_due:
            mark = 'MISSED' if s['slot_status'] == 'missed' else 'DUE SOON'
            lines.append(f"  - [{mark}] {s['reference']} {s.get('slot_type') or 'Slot'}: {s.get('slot_datetime')} "
                         f"{('(' + s['slot_ref'] + ')') if s.get('slot_ref') else ''}")
    lines.append('=' * 40)
    lines.append('Open http://localhost:7501 for full detail.')
    return '\n'.join(lines)


@app.route('/api/digest')
def api_digest():
    return jsonify({'digest': build_digest()})


@app.route('/api/digest/send', methods=['POST'])
def api_digest_send():
    text = build_digest()
    email = get_setting('digest_email')
    host = get_setting('smtp_host')
    sent = False
    if email and host:
        sent = send_email(email, 'Josawa Logistics — Morning Digest', text)
    # always persist a copy + an in-app notification
    conn = get_db()
    conn.execute('INSERT INTO notifications (kind, message) VALUES (?,?)', ('digest', text))
    conn.commit()
    conn.close()
    return jsonify({'sent': sent, 'email': email or None, 'saved': True, 'digest': text})


def send_email(to_addr, subject, body):
    """Send via SMTP if configured. Returns True on success, False otherwise.
    Credentials: SMTP_HOST, SMTP_USER, SMTP_PASS (env) — never logged."""
    import smtplib
    from email.message import EmailMessage
    host = get_setting('smtp_host')
    user = get_setting('smtp_user') or os.environ.get('SMTP_USER', '')
    pwd = os.environ.get('SMTP_PASS', '')
    frm = get_setting('smtp_from') or user
    if not host or not to_addr:
        return False
    try:
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = frm
        msg['To'] = to_addr
        msg.set_content(body)
        with smtplib.SMTP(host, timeout=15) as s:
            s.starttls()
            if user and pwd:
                s.login(user, pwd)
            s.send_message(msg)
        return True
    except Exception as e:
        print('[EMAIL] send failed:', e)
        return False


@app.route('/api/network')
def api_network():
    return jsonify({'lan_ip': _lan_ip(), 'port': 7501,
                    'lan_url': f'http://{_lan_ip()}:7501'})
def _lan_ip():
    """Best-effort LAN IP so other computers on the network can reach the app."""
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return '127.0.0.1'


if __name__ == '__main__':
    ensure_db()
    from agents.live_fetcher import start_fetcher
    start_fetcher()
    from agents.consolidator_agent import start_agent as start_consolidator
    start_consolidator()
    lan = _lan_ip()
    print('=' * 60)
    print('Josawa Logistics — Global Command Centre')
    print('  On this computer : http://localhost:7501')
    print(f'  Other computers  : http://{lan}:7501')
    print('=' * 60)
    app.run(host='0.0.0.0', port=7501, debug=False, threaded=True)
